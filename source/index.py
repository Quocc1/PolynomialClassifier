# ------------------------- Thư Viện ------------------------- #
import time
import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
import openpyxl as ox
import xlwings as xw
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import PolynomialFeatures
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error, mean_absolute_percentage_error
from sklearn.model_selection import cross_val_score
from sklearn.linear_model import LinearRegression

# ------------------------- Biến Toàn Cục ------------------------- #
df = None
df_target = None
df_scaled = None
d = None
t = None
w = None
scaler = None
test_size = None
train = None
test = None
train_time = None
test_time = None
predict = None
actual = None
result_table = None
metric_table = None
metrics = None

# ------------------------- Hàm ------------------------- #
# 1. Hàm load dữ liệu
# uploaded_file: Tập dữ liệu
# Trả về một dataframe
@st.cache_data
def LoadData(uploaded_file):
    df = pd.read_csv(uploaded_file)
    return df

# 2. Hàm tiền xử lý dữ liệu loại bỏ các outliner và giá trị bị bỏ trống và chọn Index
# data: Tập dữ liệu dùng để xử lý
# Trả về một dataframe đã được xử lý
@st.cache_data
def PreprocessingData(data):
    try:
        data.columns = data.columns.str.title()
        data = data.reset_index(drop=True)
        if "Date" in data:
            data.set_index("Date", inplace=True)

        for column in data:
            z_scores = data[column].apply(lambda x: (
                x - data[column].mean()) / data[column].std())
            outliers = z_scores[abs(z_scores) > 3].index.tolist()
            data.drop(outliers, inplace=True)

        data.dropna(inplace=True)
    except TypeError:
        pass

    return data

# 3. Hàm tạo ra ma trận giá dùng để dự đoán X và ma trận giá thực tế T
# d: Số ngày dùng để dự đoán
# t: Số ngày muốn để dự đoán
# data: Tập dữ liệu dùng để dự đoán
# column: Cột dữ liệu muốn dự đoán
# Trả về rả về ma trận X chứa giá trị để dự đoán và ma trận T chứa giá trị thực
@st.cache_data
def PrepareData(d, t, data):
    X = []
    T = []

    for i in range(len(data) - d - t + 1):
        X.append(data[i:i+d])
        T.append(data[i+d:i+d+t])

    X = np.array(X).reshape(-1).reshape(len(X), d)
    T = np.array(T).reshape(-1).reshape(len(T), t)
    return X, T

# 4. Hàm chuyển X thành dạng mở rộng đa thức theo bậc bất kỳ
# X: Ma trận chứa dữ liệu để dự đoán
# degree: Bậc của đa thức, mặc định là 2
# Trả ma trận M chứa các dữ liệu đã được biến đổi thành các đa thức.
@st.cache_data
def PolyExpansion(X, degree=2):
    poly = PolynomialFeatures(degree)
    M = poly.fit_transform(X)

    return M

# 5. Hàm tính trọng số w theo hàm loss Least Squared Error
@st.cache_data
def Model(d, t, data, degree):
    X, T = PrepareData(d, t, data)
    M = PolyExpansion(X, degree)
    s = np.dot(M.T, M)
    if np.linalg.det(s) == 0:
        s += 1e-3 * np.identity(M.shape[1])

    w = []
    for i in range(t):
        w.append(np.dot(np.dot(np.linalg.inv(s), M.T), T[:, i]))

    l1_norm = np.sum(np.abs(w))
    l2_norm = np.linalg.norm(w, ord=2)

    # w = w + 1e-3 * l1_norm + 1e-3 * l2_norm
    # w = w + 1e-3 * l2_norm

    return np.array(w)

# 6. Hàm Test
@st.cache_data
def Test(w, d, t, data, degree):
    X, T = PrepareData(d, t, data)
    result = np.dot(PolyExpansion(X, degree), w.T)
    return result, T

# 7. Hàm dự đoán


def Predict():
    pass

# 8. Hàm đánh giá độ chính xác
@st.cache_data
def Score(predict, actual):
    mae = mean_absolute_error(actual, predict)
    mse = mean_squared_error(actual, predict)
    rmse = np.sqrt(mse)
    mape = mean_absolute_percentage_error(actual, predict)

    metrics = {
        "MAE": mae,
        "MSE": mse,
        "RMSE": rmse,
        "MAPE": mape
    }

    return metrics

# 9. Hàm chuẩn hóa dữ liệu
# data: Tập dữ liệu dùng để chuẩn hóa
# strategy: Phương thức chuẩn hóa
# Trả về dataframe đã được chuẩn hóa
@st.cache_data
def standardizeData(data, strategy):
    data = data.values.reshape(-1, 1)
    if strategy == 'Z-score':
        scaler = StandardScaler()
    else:
        scaler = MinMaxScaler()

    data_scaled = scaler.fit_transform(data)
    return data_scaled

# 10. Hàm tạo tập training/testing
# data: Tập dữ liệu dùng để dự đoán
# column: Cột dữ liệu muốn dự đoán
# test_size: Tỉ lệ tập test
# Trả về mảng X_train kích thước n*d và Y_train kích thước n*t
@st.cache_data
def SplitData(data, test_size):
    train, test = train_test_split(
        data, test_size=test_size/100, shuffle=False)
    return pd.DataFrame(train), pd.DataFrame(test)

# 11. Hàm rescale dữ liệu
# data: Dữ liệu muốn rescale về giá trị ban đầu
# strategy: Phương thức chuẩn hóa
# Trả về dữ liệu đã được rescale về giá trị ban đầu
@st.cache_data
def RescaleModel(data, strategy):
    if strategy == 'Z-score':
        scaler = StandardScaler()
    else:
        scaler = MinMaxScaler()

    scaler.fit(df_target.values.reshape(-1, 1))
    data_inverse = scaler.inverse_transform(data)
    return data_inverse

# 12. Hàm tính RMSE kiểm định chéo và lưu kết quả vào tệp Excel
def cv_rmse(data, degree):
    X, T = PrepareData(d, t, data)
    M = PolyExpansion(X, degree)

    # Tạo mô hình Linear Regression
    model = LinearRegression()

    # Thực hiện cross-validation và tính RMSE
    rmse_scores = np.sqrt(-cross_val_score(model, M, T, scoring="neg_mean_squared_error", cv=5))

    # Tính giá trị RMSE trung bình từ cross-validation
    mean_rmse = np.mean(rmse_scores)
    
    return mean_rmse


def FormatExcel():
    wb = ox.load_workbook('../report/output.xlsx')
    ws = wb['Sheet_1']

    for letter in ['B', 'C']:
        max_width = 0
        for row_number in range(1, ws.max_row + 1):
            width = len(ws[f'{letter}{row_number}'].value)

            if width > max_width:
                max_width = width

        ws.column_dimensions[letter].width = max_width + 1
        

    for letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
        max_width = 0
        for row_number in range(1, 3):

            width = len(str(ws[f'{letter}{row_number}'].value))
            if width > max_width:
                max_width = width

        ws.column_dimensions[letter].width = max_width + 1
    wb.save('../report/output.xlsx')

# 13. Hàm xuất kết quả ra file CSV
def SaveResult(result_table, metric_table):
    wb_name = "../report/output.xlsx"

    try:
        with pd.ExcelWriter(wb_name, engine='openpyxl') as writer:
            result_table.to_excel(writer, sheet_name="Sheet_1")
            metric_table.to_excel(writer, sheet_name="Sheet_1", startcol=5)
            st.write("**Kết quả đã được lưu trong report/output.xlsx**")
    
        FormatExcel()
    except PermissionError:
        st.write('**Không thể lưu, vui lòng đóng ứng dụng Excel đang mở**')
    

# Xóa dữ liệu lưu trong streamlit
def ClearCache():
    st.session_state.clear()


# ------------------------- Giao Diện ------------------------- #
st.title('Dự báo chứng khoán bằng Polynomial Classifier')

# Chọn tập dữ liệu
uploaded_file = st.file_uploader("**Chọn tệp dữ liệu CSV:**", type=["csv"], on_change=ClearCache)

if uploaded_file is not None:
    file_name = uploaded_file.name
    df = LoadData(uploaded_file)

    # Thông tin tập dữ liệu
    st.subheader('Thông tin tập dữ liệu ' + file_name)
    st.write(df)


# Tiền xử lý dữ liệu
if df is not None:
    with st.spinner('Đang tiền xử lý tập dữ liệu, xin vui lòng đợi.'):
        df = PreprocessingData(df)

    # Vẽ biểu đồ đường cho tập dữ liệu
    st.subheader('Biểu đồ đường tập dữ liệu ' + file_name)

    column_names = df.columns.tolist()
    selected_column_name = st.selectbox("**Chọn cột:**", column_names)
    st.line_chart(data=df, y=selected_column_name)

    # Chọn tham số
    st.subheader('Tạo mô hình dự đoán trên tập dữ liệu ' + file_name)

    # Chọn cột để dự đoán
    selected_predict_column_name = st.selectbox(
        "**Chọn cột để dự đoán:**", column_names, on_change=ClearCache)

    df_target = df[selected_column_name]

    # Chọn ngày để dự đoán
    col11, col21 = st.columns(2)
    with col11:
        d = st.number_input('**Số ngày dùng để dự đoán:**',
                            value=2, step=1, min_value=1, on_change=ClearCache)

    with col21:
        t = st.number_input('**Số ngày muốn dự đoán:**', value=1,
                            step=1, min_value=1, on_change=ClearCache)

    # Chọn tỉ lệ chia tập train/test
    col12, col22 = st.columns(2)
    with col12:
        train_size = st.number_input(
            '**Chia tỉ lệ training %:**', min_value=1, max_value=99, value=70, step=1, on_change=ClearCache)
        test_size = 100 - train_size
        st.write('**Tỷ lệ testing: **' + str(test_size) + '**%**')

    # Chọn bậc của đa thức
    with col22:
        degree = st.number_input(
            '**Số bậc của đa thức:**', value=2, step=1, min_value=1, on_change=ClearCache)

    # Chuẩn hóa dữ liệu
    col13, col23 = st.columns(2)
    with col13:
        st.subheader('Chuẩn hóa dữ liệu')
        standardize_strategy = st.radio(
            "**Chọn phương thức chuẩn hóa:**",
            ["Z-score", "Min-max"], horizontal=True)

        if st.button("Chuẩn hóa dữ liệu", type="primary", on_click=ClearCache):
            df_scaled = standardizeData(df_target, standardize_strategy)
            st.session_state.df_scaled = df_scaled

    # Hiệu chỉnh dữ liệu
    # with col23:
    #     st.subheader('Hiệu chỉnh dữ liệu')
    #     regularize_strategy = st.radio(
    #         "Chọn phương thức hiệu chỉnh: ",
    #         ["L1", "L2", "L1 + L2"], horizontal=True)

    #     if st.button("Hiệu chỉnh dữ liệu", type="primary", on_click=ClearCache):
    #         # df_scaled = regularizeData(df_target, regularize_strategy)
    #         # st.session_state.df_scaled = df_scaled
    #         pass

    # In ra 5 dòng dữ liệu trước và sau chuẩn hóa
    col14, col24 = st.columns(2)
    with col14:
        st.write('**Dữ liệu trước khi chuẩn hóa:**')
        st.write(df_target[:5])
    with col24:
        st.write('**Dữ liệu sau khi chuẩn hóa:**')
        try:
            df_scaled = st.session_state.df_scaled
            st.write(df_scaled[:5])
        except AttributeError:
            df_scaled = None

    # Chia tập training/testing
    train, test = SplitData(
        df_target if df_scaled is None else df_scaled, test_size)
    st.write('**Kích thước tập train:**', train.size,
             '**Kích thước tập test:**', test.size)

    # Training
    if st.button('Train Mô Hình {}'.format('Chưa Chuẩn Hóa' if df_scaled is None else 'Đã Chuẩn Hóa'), type='primary'):
        with st.spinner('Mô hình đang được train, xin vui lòng đợi.'):
            start_time = time.time()
            w = Model(d, t, train, degree)
            train_time = "{:.2f}".format(time.time() - start_time)
            
            st.session_state.train_time = train_time
            st.session_state.w = w

    # In ma trận w ra màn hình
    try:
        w = st.session_state.w
        st.write('**Ma trận trọng số có dạng:**')
        vmatrix_latex = "\omega = \\begin{vmatrix} "

        for i in range(len(w)):
            vmatrix_latex += " & ".join(["{:.4f}".format(x)
                                        for x in w[i]]) + " \\\\ "
        vmatrix_latex += " \\end{vmatrix}"

        st.latex(vmatrix_latex)
    except AttributeError:
        w = None

    # Testing
    if st.button("Test Mô Hình", type="primary"):
        try:
            w = st.session_state.w
            try:
                start_time = time.time()
                predict, actual = Test(w, d, t, test, degree)
                test_time = "{:.2f}".format(time.time() - start_time)

                st.session_state.test_time = test_time
                st.session_state.predict = predict
                st.session_state.actual = actual
            except ValueError:
                st.write('**Tập dữ liệu quá nhỏ, không thể test**')
        except AttributeError:
            st.write('**Vui lòng train tập dữ liệu trước khi test**')


    # In 5 dòng đầu của dự đoán và thực tế
    try:
        predict = st.session_state.predict
        actual = st.session_state.actual
        predict_list = predict.tolist()
        actual_list = actual.tolist()
        
        result_table = pd.DataFrame(
            {"Dự đoán": predict_list, "Thực tế": actual_list})
        
        st.session_state.result_table = result_table
        
        st.table(result_table[:5])

    except AttributeError:
        predict = None
        actual = None

    # Tính lỗi và thời gian thực thi
    if predict is not None:
        st.subheader('Các thông số đánh giá')
        
        metrics = Score(predict, actual)
        metrics.update({"Thời gian train": st.session_state.train_time, "Thời gian test": st.session_state.test_time})
        metric_table = pd.DataFrame(metrics, index=["PC"])
        st.session_state.metric_table = metric_table

        st.table(metric_table)

        # Lưu kết quả ra file output.xlsx trong thư mục result
        if result_table is not None:
            if st.button("Xuất kết quả ra file CSV", type="primary"):
                SaveResult(st.session_state.result_table, st.session_state.metric_table)


        # Vẽ biểu đồ đường dự báo và thực tế
        st.subheader('Biểu đồ so sánh')

        index_column = np.array(df_target[train.size + d:].index).flatten()

        predict = pd.DataFrame({'Dự đoán': predict.flatten()})
        predict = predict.set_index(index_column)
        actual = pd.DataFrame({'Thực tế': actual.flatten()})
        actual = actual.set_index(index_column)

        df_result = pd.concat([predict, actual], axis=1)

        fig1 = px.line(df_result, x=df_result.index, y=df_result.columns,
                        color_discrete_sequence=["red", "blue"])
        fig1.update_traces(patch={"line": {"width": 1, "dash": 'dot'}})

        st.plotly_chart(fig1)

        # Rescale model và vẽ lại biểu đồ
        if predict is not None and df_scaled is not None:
            predict_inv = RescaleModel(predict, standardize_strategy)
            actual_inv = RescaleModel(actual, standardize_strategy)

            predict_inv = pd.DataFrame({'Dự đoán': predict_inv.flatten()})
            predict_inv = predict_inv.set_index(index_column)
            actual_inv = pd.DataFrame({'Thực tế': actual_inv.flatten()})
            actual_inv = actual_inv.set_index(index_column)

            df_rescale = pd.concat([predict_inv, actual_inv], axis=1)

            if st.button('Rescale Mô Hình', type='primary'):
                fig2 = px.line(df_rescale, x=df_rescale.index,
                            y=df_rescale.columns, color_discrete_sequence=["red", "blue"])
                fig2.update_traces(patch={"line": {"width": 1, "dash": 'dash'}})

                st.plotly_chart(fig2)

# Áp dụng mô hình cho tập dữ liệu khác

# df2 = pd.read_csv('./data/XRX.csv')
# df2_index = df2['Date']
# df2 = df2.set_index(df2_index)
# st.write('Ma trận trọng số có dạng:')
# vmatrix_latex = "\omega = \\begin{vmatrix} "

# for i in range(len(w)):
#     vmatrix_latex += " & ".join(["{:.4f}".format(x) for x in w[i]]) + " \\\\ "
# vmatrix_latex += " \\end{vmatrix}"

# st.latex(vmatrix_latex)
# predict2, actual2 = Test(w, d, t, df2[selected_column_name], 2)
# predict2 = pd.DataFrame({'Dự đoán': predict2.flatten()})
# predict2 = predict2.set_index(df2_index[d:])
# actual2 = pd.DataFrame({'Thực tế': actual2.flatten()})
# actual2 = actual2.set_index(df2_index[d:])
# df_another = pd.concat([predict2, actual2], axis=1)

# fig3 = px.line(df_another, x=df_another.index, y=df_another.columns, color_discrete_sequence=["red", "blue"])
# fig3.update_traces(patch={"line": {"width": 1, "dash": 'dash'}})

# st.plotly_chart(fig3)
# So sánh mô hình với các bậc đa thức khác

# Nhập dữ liệu để dự đoán
